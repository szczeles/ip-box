#!/usr/bin/env python3

import argparse
import os
import csv
import dateutil.parser
from decimal import Decimal
import yaml

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter


class KPiRRow:

    def __init__(self, row):
        if not len(row) == 16:
            raise ValueError("wrong KPiR row")
        self._row = row

    @property
    def number(self):
        return self._row[0]

    @property
    def date(self):
        return dateutil.parser.isoparse(self._row[1])

    @property
    def invoice_number(self):
        return self._row[2]

    @property
    def company_name(self):
        return self._row[3]

    @property
    def company_address(self):
        return self._row[4]

    @property
    def description(self):
        return self._row[5]

    @property
    def sales_income(self):
        return as_decimal(self._row[6])

    @property
    def other_income(self):
        return as_decimal(self._row[7])

    @property
    def income(self):
        return as_decimal(self._row[8])

    @property
    def is_income(self):
        return self.income > 0

    @property
    def other_cost(self):
        return as_decimal(self._row[12])

    @property
    def cost(self):
        return as_decimal(self._row[13])

    @property
    def is_cost(self):
        return self.cost > 0


def as_decimal(value):
    return Decimal(value.replace(",", "."))


class KPiR:

    def __init__(self, year, rows):
        self._year = year
        self._rows = [KPiRRow(row) for row in rows]

    @property
    def costs(self):
        return self.filter(lambda row: row.is_cost)

    @property
    def incomes(self):
        return self.filter(lambda row: row.is_income)

    def filter(self, condition):
        return [row for row in self._rows if condition(row)]

    def filter_year(self, year):
        return self.filter(lambda row: row.date.year == year)

    @property
    def rows(self):
        return self._rows

    def compute_totals(self, year):
        incomes = Decimal("0")
        costs = Decimal("0")
        for row in self.filter_year(year):
            if row.is_cost:
                costs += row.cost
            if row.is_income:
                incomes += row.income
        return incomes, costs


class KPiRCsvReader:

    def __init__(self, file_path, year, with_header=True):
        self._file_path = file_path
        self._year = str(year)
        self._with_header = with_header

    def __enter__(self):
        self._file = open(self._file_path, 'r')
        self._reader = csv.reader(self._file)
        return self

    def __exit__(self, *args):
        self._file.close()
        return True

    def _is_applicable(self, row):
        return len(row) == 16 and row[1].startswith(self._year)

    def _read_rows(self):
        row_number = 0
        for row in self._reader:
            row_number += 1
            if row_number == 1 and self._with_header:
                continue
            if self._is_applicable(row):
                yield row

    def read(self):
        return KPiR(self._year, [row for row in self._read_rows()])


class Project:

    def __init__(self, id, descriptor, common):
        self._id = id
        self._descriptor = descriptor
        self._common = common

    @property
    def kpwi_id(self):
        return self._id

    @property
    def name(self):
        return self._get_property('name')

    @property
    def kind(self):
        return self._get_property('kind', required=False, default='Autorskie prawo do programu komputerowego')

    @property
    def description(self):
        return self._get_property('description')

    @property
    def interpretation_number(self):
        return self._get_property('interpretationNumber')

    @property
    def start_date(self):
        return self._get_property('startDate')

    @property
    def end_date(self):
        return self._get_property('endDate', required=False)

    @property
    def results(self):
        return self._get_property('results')

    @property
    def employees(self):
        return ['Michał Lula']

    def find_costs(self, kpir):
        return kpir.filter(lambda row: row.description in self._descriptor['description'])

    def is_related(self, row):
        return row.description in self._descriptor['costs']['description']

    def _get_property(self, name, required=True, default=None):
        if name in self._descriptor:
            return self._descriptor[name]
        elif name in self._common:
            return self._common[name]
        elif required:
            raise ValueError(f"missing property '{name}' for project {self._id} ({self._get_property('name', False)})")
        else:
            return default


class ProjectsDescriptor:

    def __init__(self, file_path):
        self._file_path = file_path

    def __enter__(self):
        self._file = open(self._file_path)
        return self

    def load(self):
        projects = yaml.safe_load(self._file)
        if 'projects' not in projects:
            raise ValueError("projects descriptor must define 'projects' property")
        general = projects['general'] if 'general' in projects else {}
        return [Project(f'KPWI-{idx+1:03d}', project, general) for idx, project in enumerate(projects['projects'])]

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._file.close()
        if exc_type is not None:
            print(f"{exc_type}, {exc_val}, {exc_tb}")
            return False
        return True


class SalesRecordsReport:

    _headers=dict(
        projects = dict(
            title='Projekty',
            header={
                'A1': 'Projekt',
                'B1': 'Identyfikator KPWI',
                'C1': 'Rodzaj',
                'D1': 'Numer Interpretacji Indywidualnej',
                'E1': 'Opis Projektu',
                'F1': 'Data Rozpoczęcia',
                'G1': 'Data Zakończenia',
                'H1': 'Prace Stworzone',
                'I1': 'Pracownicy zaangażowany w projekt'
            }),
        sales_records = dict(
            title='Dokumenty księgowe',
            header={
                'A1:A2': 'Data zdarzenia',
                'B1:B2': 'Nr dowodu księgowego',
                'C1:C2': 'Numer w KPiR',
                'D1:D2': 'Opis zdarzenia gospodarczego',
                'E1:E2': 'Identyfikator KPWI',
                'F1:H1': 'Przychód',
                'F2': 'Zbycie KPWI',
                'G2': 'Pozostałe',
                'H2': 'SUMA',
                'I1:M1': 'Koszty działalności badawczo - rozwojowej',
                'I2': 'Kategoria A',
                'J2': 'Kategoria B',
                'K2': 'Kategoria C',
                'L2': 'Kategoria D',
                'M2': 'SUMA'
            }),
        results = dict(
            title='Wyniki',
            header={
                'A1': 'Identyfikator KPWI',
                'B1': 'Pole PIT IP',
                'C1': 'Wartość',
                'A2': 'Przychody z kwalifikowanych praw',
                'B2': '16',
                'A3': 'Koszty Kwalifikowane',
                'A4': 'Koszty pośrednie',
                'A5': 'Koszty uzyskania przychodu związane z kwalifikowanymi prawami',
                'B5': '17',
                'A6': 'Nexus niezaokrąglony',
                'A7': 'Nexus',
                'A8': 'Kwalifikowany dochód (dochody) z kwalifikowanych praw obliczony (obliczone) zgodnie z art. 30ca ust. 4 ustawy',
                'B8': '18 / 42',
                'A9': 'Dochód z kwalifikowanych praw niepodlegający opodatkowaniu stawką 5%, o której mowa w art. 30ca ust. 1 ustawy',
                'B9': '19 / 20 (jeżeli ujemny)',
                'A10': 'Podstawa opodatkowania',
                'B10': '31',
                'A11': 'Podstawa opodatkowania',
                'B11': '42',
                'A12': 'Podatek',
                'B12': '44',
                'A13': 'Podatek',
                'B13': '47',
            })
    )

    def __init__(self, kpir, projects):
        self._kpir = kpir
        self._projects = projects

    def generate(self, year, path, report_name):
        projects = self._get_projects(year)
        wb = Workbook()
        self._generate_projects(wb.active, projects)
        self._generate_records(wb.create_sheet(), year, projects)
        self._generate_results(wb.create_sheet(), year, projects)
        print(f"saving to {path}/{report_name}")
        wb.save(filename=f'{path}/{report_name}')

    def _get_projects(self, year):
        return [project for project in self._projects if project.start_date.year <= year <= project.end_date.year]

    def _generate_header(self, sheet, header):
        sheet.title = header['title']
        for (cell, value) in header['header'].items():
            if ':' in cell:
                self._set_header(sheet, cell.partition(":")[0], value)
                sheet.merge_cells(cell)
            else:
                self._set_header(sheet, cell, value)

    def _set_header(self, sheet, cell, value):
        sheet[cell] = value
        sheet[cell].font = Font(color='FF000000', bold=True)
        sheet[cell].fill = PatternFill("solid", fgColor=Color(indexed=22))

    def _generate_projects(self, sheet, projects):
        self._generate_header(sheet, self._headers['projects'])
        index = 2
        for project in projects:
            sheet[f'A{index}'] = project.name
            sheet[f'B{index}'] = project.kpwi_id
            sheet[f'C{index}'] = project.kind
            sheet[f'D{index}'] = project.interpretation_number
            sheet[f'E{index}'] = project.description
            sheet[f'F{index}'] = project.start_date.isoformat()
            sheet[f'G{index}'] = project.end_date.isoformat() if project.end_date else ""
            sheet[f'H{index}'] = ", ".join(project.results)
            sheet[f'I{index}'] = ", ".join(project.employees)
            index += 1

    def _generate_records(self, sheet, year, projects):
        self._generate_header(sheet, self._headers['sales_records'])
        index = 3
        for record in self._get_records(year, projects):
            sheet[f'A{index}'] = record.date
            sheet[f'B{index}'] = record.invoice_number
            sheet[f'C{index}'] = record.kpir_number
            sheet[f'D{index}'] = record.description
            sheet[f'E{index}'] = ", ".join(record.projects_kpwi_ids)
            sheet[f'F{index}'] = record.income_kpwi
            sheet[f'G{index}'] = record.income_other
            sheet[f'H{index}'] = f'=F{index}+G{index}'
            sheet[f'I{index}'] = record.cost_a
            sheet[f'J{index}'] = record.cost_b
            sheet[f'K{index}'] = record.cost_c
            sheet[f'L{index}'] = record.cost_d
            sheet[f'M{index}'] = f'=I{index}+J{index}+K{index}+L{index}'
            index += 1

    def _get_records(self, year, projects):
        for row in self._kpir.filter_year(year):
            related_projects = []
            for project in projects:
                if project.is_related(row):
                    related_projects.append(project)
            if related_projects:
                yield KPWIRecord(row, related_projects)

    def _generate_results(self, sheet, year, projects):
        self._generate_header(sheet, self._headers['results'])
        total_incomes, total_costs = self._kpir.compute_totals(year)
        index = 4
        min_column = None
        max_column = None
        for project in projects:
            column = get_column_letter(index)
            self._set_header(sheet, f'{column}1', project.kpwi_id)
            sheet[f'{column}2'] = f"={self._sales_record_for_kpwi(column + '1', 'F')}"
            sheet[f'{column}3'] = f"={self._sales_record_for_kpwi(column + '1', 'M')}"
            sheet[f'{column}4'] = f"=({column}2/{total_incomes})*({total_costs}-C3)"
            sheet[f'{column}5'] = f'={column}3 + {column}4'
            sheet[f'{column}6'] = self._nexus_formula(column + '1')
            sheet[f'{column}7'] = f'=IF({column}6>1,1,{column}6)'
            sheet[f'{column}8'] = f'=({column}2-{column}5)*{column}7'
            sheet[f'{column}9'] = f'=({column}2-{column}5)*(1-{column}7)'
            sheet[f'{column}10'] = f'={column}8'
            sheet[f'{column}11'] = f'={column}8'
            sheet[f'{column}12'] = f'={column}8*0.05'
            sheet[f'{column}13'] = f'={column}8*0.05'
            if min_column is None:
                min_column = column
            max_column = column
            index += 1
        for row in [2, 3, 4, 5, 8, 9, 10, 11, 12, 13]:
            sheet[f'C{row}'] = f'=SUM({min_column}{row}:{max_column}{row})'

    def _nexus_formula(self, kpwi_cell):
        nominator = '+'.join([self._sales_record_for_kpwi(kpwi_cell, cost_column) for cost_column in ['I', 'J', 'K']])
        return f"=1.3*({nominator})/{self._sales_record_for_kpwi(kpwi_cell, 'M')}"

    def _sales_record_for_kpwi(self, kpwi_cell, column):
        sales_records_tab = self._headers['sales_records']['title']
        return f"SUMIF('{sales_records_tab}'!E:E,{kpwi_cell},'{sales_records_tab}'!{column}:{column})"


class KPWIRecord:

    def __init__(self, row, projects, kpwi_rate=Decimal("0.8")):
        self._row = row
        self._projects = projects
        self._kpwi_rate = kpwi_rate

    @property
    def date(self):
        return self._row.date.isoformat()

    @property
    def invoice_number(self):
        return self._row.invoice_number

    @property
    def kpir_number(self):
        return self._row.number

    @property
    def description(self):
        return self._row.description

    @property
    def projects_kpwi_ids(self):
        return [project.kpwi_id for project in self._projects]

    @property
    def income_kpwi(self):
        return self._kpwi_rate * (self._row.income if self._row.is_income else 0)

    @property
    def income_other(self):
        return (1 - self._kpwi_rate) * (self._row.income if self._row.is_income else 0)

    @property
    def income_total(self):
        return self.income_kpwi + self.income_other

    @property
    def cost_a(self):
        return self._row.cost if self._row.is_cost else 0

    @property
    def cost_b(self):
        return 0

    @property
    def cost_c(self):
        return 0

    @property
    def cost_d(self):
        return 0

    @property
    def cost_total(self):
        return self.cost_a + self.cost_b + self.cost_c + self.cost_d


if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument('--year', help='report year', type=int, required=True)
    parser.add_argument('--kpir', help='KPiR CSV file path', required=True)
    parser.add_argument('--projects', help="path to projects descriptor file", required=True)
    parser.add_argument('--path', help="Reports path", required=False, default="out")
    parser.add_argument('--report-name', help="Report file name", required=False, default='Ewidencja-IP-Box.xlsx')

    args = parser.parse_args()

    report_path = f'{args.path}/{args.year}'

    os.makedirs(report_path, exist_ok=True)

    with KPiRCsvReader(args.kpir, args.year) as reader:
        with ProjectsDescriptor(args.projects) as descriptor:
            SalesRecordsReport(reader.read(), descriptor.load()).generate(args.year, report_path, args.report_name)


