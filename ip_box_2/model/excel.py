import typing

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color


class BaseSheet:

    _title = ''
    _header = None
    _header_font = Font(color='FF000000', bold=True)
    _header_fill = PatternFill("solid", fgColor=Color(indexed=22))
    _columns_width = None
    _wrap = False

    def __init__(self, sheet, title = None, header : typing.Mapping[str, str] = None, header_font=None, header_fill=None, column_width = None, **kwargs):
        self._sheet = sheet
        if title:
            self._title = title
        sheet.title = self._title
        if header_font:
            self._header_font = header_font
        if header_fill:
            self._header_fill = header_fill
        if header:
            self._header = header
        for cell, cell_title in (self._header or {}).items():
            self.set_header(cell, cell_title)
        if column_width:
            self._columns_width = column_width
        for cell, width in (self._columns_width or {}).items():
            sheet.column_dimensions[cell].width = width
        self._wrap = kwargs.get('wrap', self._wrap)

    @property
    def sheet(self):
        return self._sheet

    def set_header(self, cell, value):
        self.set_value(cell, value, font=self._header_font, fill=self._header_fill)

    def set_value(self, cell, value, font=None, fill=None, **kwargs):
        first_cell = cell.partition(":")[0] if ':' in cell else cell
        self._sheet[first_cell] = value
        if font:
            self[first_cell].font = font
        if fill:
            self[first_cell].fill = fill
        if kwargs.get('wrap', self._wrap):
            self[first_cell].alignment = Alignment(wrap_text=True)
        if ':' in cell:
            self._sheet.merge_cells(cell)

    def __setitem__(self, key, value):
        self.set_value(key, value)

    def __getitem__(self, item):
        return self._sheet[item]