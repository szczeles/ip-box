import os

import click
import datetime
import decimal
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from context import pass_ip_box, IpBoxContext, Project
from model.excel import BaseSheet
from model.kpir import KpirCsvReader
from model.classification import ClassifiedKpirRow
from common import CsvReader, AggregatedTimeEntry


class WagedIpIncome:

    def __init__(self, wage, ip_hours, other_hours):
        self._wage = wage
        self._ip_hours = ip_hours
        self._other_hours = other_hours

    @property
    def ip_hours(self):
        return self._ip_hours

    @property
    def ip_income(self):
        return self._wage * self._ip_hours

    @property
    def other_hours(self):
        return self._other_hours

    @property
    def other_income(self):
        return self._wage * self._other_hours

    @property
    def wage(self):
        return self._wage


class IpBoxIncomeRegistry:

    def __init__(self, context: IpBoxContext):
        self._context = context
        self._ip_income_registry = {}
        self._total_income_registry = {}

    def add_ip_income(self, year, month, project, ip_hours, other_hours):
        if year not in self._ip_income_registry:
            self._ip_income_registry[year] = {}
        year_registry = self._ip_income_registry[year]
        if month not in year_registry:
            year_registry[month] = {}
        month_registry = year_registry[month]
        if project.id in month_registry:
            raise ValueError(f'duplicated income for {year}-{month:02d} - {project.id}')
        wage = self._context.get_wage(datetime.date(year, month, 1))
        month_registry[project.id] = WagedIpIncome(wage, ip_hours, other_hours)

    def add_income(self, year, month, income):
        if year not in self._total_income_registry:
            self._total_income_registry[year] = {}
        year_registry = self._total_income_registry[year]
        if month not in year_registry:
            year_registry[month] = 0
        year_registry[month] += income

    def get_ip_income(self, year, month, project_id) -> WagedIpIncome:
        if year not in self._ip_income_registry:
            raise ValueError(f'missing date for year {year}')
        if month not in self._ip_income_registry[year]:
            raise ValueError(f'missing date for month {year}-{month:02d}')
        if project_id not in self._ip_income_registry[year][month]:
            raise ValueError(f'missing data for project {year}-{month:02d} - {project_id}')
        return self._ip_income_registry[year][month][project_id]

    def get_total_income(self, year, month):
        if year not in self._total_income_registry:
            raise ValueError(f'missing date for year {year}')
        if month not in self._total_income_registry[year]:
            raise ValueError(f'missing date for month {year}-{month:02d}')
        return self._total_income_registry[year][month]

    def get_income_ratio(self, year, month, project_id):
        ip_income = self.get_ip_income(year, month, project_id)
        total_income = self.get_total_income(year, month)
        return decimal.Decimal(ip_income.ip_income) / total_income


class ProjectsSummarySheet(BaseSheet):
    _title = 'Zestawienie projektów'
    _header = {
        'A1': 'Projekt',
        'B1': 'Identyfikator KPWI',
        'C1': 'Rodzaj',
        'D1': 'Numer Interpretacji Indywidualnej',
        'E1': 'Opis Projektu',
        'F1': 'Data Rozpoczęcia',
        'G1': 'Data Zakończenia',
        'H1': 'Prace Stworzone',
        'I1': 'Pracownicy zaangażowany w projekt'
    }
    _columns_width = {
        'A': 50,
        'B': 20,
        'C': 30,
        'D': 20,
        'E': 50,
        'F': 15,
        'G': 15,
        'H': 50,
        'I': 20
    }
    _wrap = True

    def __init__(self, sheet, projects, **kwargs):
        super().__init__(sheet, **kwargs)
        self._projects = projects
        for idx, project in enumerate(projects):
            row = idx + 2
            for (letter, value) in (
                ('A', project.name),
                ('B', project.id),
                ('C', 'Autorskie prawo do programu komputerowego'),
                ('D', project.interpretation_number),
                ('E', project.description),
                ('F', project.start_date.isoformat()),
                ('G', project.end_date.isoformat() if project.end_date else ''),
                ('H', ', '.join([result.name for result in project.results])),
                ('I', project.employee)
            ):
                self[f'{letter}{row}'] = value


class SalesRecordsSummarySheet(BaseSheet):
    _title = 'Zestawienie ksiąg'
    _header = {
        'A1:A2': 'Data zdarzenia',
        'B1:B2': 'Nr dowodu księgowego',
        'C1:C2': 'Numer w KPiR',
        'D1:D2': 'Opis zdarzenia gospodarczego',
        'E1:E2': 'Identyfikator KPWI',
        'F1:H1': 'Przychód',
        'F2': 'Zbycie KPWI',
        'G2': 'Pozostałe',
        'H2': 'SUMA',
        'I1:M1': 'Koszty działalności badawczo - rozwojowej',
        'I2': 'Kategoria A',
        'J2': 'Kategoria B',
        'K2': 'Kategoria C',
        'L2': 'Kategoria D',
        'M2': 'SUMA',
        'N2': 'Proporcja przychodu z KPiW',
        'O2': 'Koszt orginalny',
        'P2': 'Przychód orginalny',
        'R2': 'Liczba godzin IP',
        'S2': 'Liczba godzin pozstałych',
        'T2': 'Stawka godzinowa'
    }
    _columns_width = {
        'A': 10,
        'B': 20,
        'C': 5,
        'D': 40,
        'E': 20,
        'F': 10,
        'G': 10,
        'H': 10,
        'I': 10,
        'J': 10,
        'K': 10,
        'L': 10,
        'M': 10,
        'N': 10,
        'O': 10,
        'P': 10,
        'R': 10,
        'S': 10,
        'T': 10
    }

    def __init__(self, sheet, income_registry: IpBoxIncomeRegistry, **kwargs):
        super().__init__(sheet, self._title, **kwargs)
        self._row_idx = 0
        self._income_registry = income_registry

    def add(self, row: ClassifiedKpirRow):
        if row.is_kpiw:
            for project_id in row.projects_ids:
                self._row_idx += 1
                idx = self._row_idx + 2
                ratio = self._income_registry.get_income_ratio(row.date.year, row.date.month, project_id)
                if row.is_income:
                    income = self._income_registry.get_ip_income(row.date.year, row.date.month, project_id)
                    ip_income = income.ip_income
                    other_income = income.other_income
                    self[f'P{idx}'] = row.income
                    self[f'R{idx}'] = income.ip_hours
                    self[f'S{idx}'] = income.other_hours
                    self[f'T{idx}'] = income.wage
                else:
                    ip_income = 0
                    other_income = 0
                    self[f'O{idx}'] = row.cost
                self[f'A{idx}'] = row.date.isoformat()
                self[f'B{idx}'] = row.invoice_number
                self[f'C{idx}'] = row.number
                self[f'D{idx}'] = row.description
                self[f'E{idx}'] = project_id
                self[f'F{idx}'] = ip_income
                self[f'G{idx}'] = other_income
                self[f'H{idx}'] = f'=F{idx}+G{idx}'
                self[f'I{idx}'] = ratio * row.cost_a
                self[f'J{idx}'] = ratio * row.cost_b
                self[f'K{idx}'] = ratio * row.cost_c
                self[f'L{idx}'] = ratio * row.cost_d
                self[f'M{idx}'] = f'=I{idx}+J{idx}+K{idx}+L{idx}'
                self[f'N{idx}'] = ratio
                self[f'N{idx}'].number_format = '0.00%'


class ResultsSheet(BaseSheet):
    _title = 'Wyniki'

    _header = {
        'A1': 'Identyfikator KPWI',
        'B1': 'Pole PIT IP',
        'C1': 'Wartość',
        'A2': 'Przychody z kwalifikowanych praw',
        'B2': '16',
        'A3': 'Koszty Kwalifikowane',
        'A4': 'Koszty pośrednie',
        'A5': 'Koszty uzyskania przychodu związane z kwalifikowanymi prawami',
        'B5': '17',
        'A6': 'Nexus niezaokrąglony',
        'A7': 'Nexus',
        'A8': 'Kwalifikowany dochód z kwalifikowanych praw obliczony zgodnie z art. 30ca ust. 4 ustawy',
        'B8': '18 / 42',
        'A9': 'Dochód z kwalifikowanych praw niepodlegający opodatkowaniu stawką 5%, o której mowa w art. 30ca ust. 1 ustawy',
        'B9': '19 / 20 (jeżeli ujemny)',
        'A10': 'Podstawa opodatkowania',
        'A14': 'Łączny Przychód',
        'A15': 'Łączne koszty',
        'A16': 'Łączny dochód',
        'A17': 'Odliczenia od dochodu',
        'A18': 'Dochód po odczliczeniach',
        'A19': 'Dochód nie związany z KPWI',
        'A20': 'Podatek nie związany z KPWI',
        'A21': 'Podatek Łącznie',
        'A22': 'Podatek Pierwotny',
        'A23': 'Nadpłata',
        'B10': '31',
        'A11': 'Podstawa opodatkowania',
        'B11': '42',
        'A12': 'Podatek',
        'B12': '44',
        'A13': 'Podatek',
        'B13': '47',
    }
    _kpiw_column_width = 20

    _columns_width = {
        'A': 50,
        'B': 50,
        'C': 20,
    }

    def __init__(self, sheet, projects, **kwargs):
        super().__init__(sheet, **kwargs)
        self._projects = projects
        self._total_incomes = 0
        self._total_costs = 0
        self._total_income_reduction = 0

    def add(self, row: ClassifiedKpirRow):
        if row.is_income:
            self._total_incomes += row.income
        elif row.is_cost:
            self._total_costs += row.cost
        else:
            raise ValueError(f'strange row: {row}')

    def flush(self):
        index = 4
        min_column = None
        max_column = None
        self['C14'] = self._total_incomes
        self['C15'] = self._total_costs
        self['C16'] = '=C14-C15'
        self['C17'] = self._total_income_reduction
        self['C18'] = '=C16-C17'
        self['C19'] = '=C18-C8'
        self['C20'] = '=C19*0.19'
        self['C21'] = '=C13+C20'
        self['C22'] = '=C18*0.19'
        self['C23'] = '=C22-C21'
        for project in self._projects:
            column = get_column_letter(index)
            self.set_header(f'{column}1', project.id)
            self.sheet.column_dimensions[column].width = self._kpiw_column_width
            self[f'{column}2'] = f"={self._sales_record_for_kpwi(column + '1', 'F')}"
            self[f'{column}3'] = f"={self._sales_record_for_kpwi(column + '1', 'M')}"
            self[f'{column}4'] = f"=({column}2/C14)*(C15-C3)"
            self[f'{column}5'] = f'={column}3 + {column}4'
            self[f'{column}6'] = self._nexus_formula(column + '1')
            self[f'{column}7'] = f'=IF({column}6>1,1,{column}6)'
            self[f'{column}8'] = f'=({column}2-{column}5)*{column}7'
            self[f'{column}9'] = f'=({column}2-{column}5)*(1-{column}7)'
            self[f'{column}10'] = f'={column}8'
            self[f'{column}11'] = f'={column}8'
            self[f'{column}12'] = f'={column}8*0.05'
            self[f'{column}13'] = f'={column}8*0.05'
            if min_column is None:
                min_column = column
            max_column = column
            index += 1
        for row in [2, 3, 4, 5, 8, 9, 10, 11, 12, 13]:
            self[f'C{row}'] = f'=SUM({min_column}{row}:{max_column}{row})'

    def _nexus_formula(self, kpwi_cell):
        nominator = '+'.join([self._sales_record_for_kpwi(kpwi_cell, cost_column) for cost_column in ['I', 'J', 'K']])
        return f"=1.3*({nominator})/{self._sales_record_for_kpwi(kpwi_cell, 'M')}"

    def _sales_record_for_kpwi(self, kpwi_cell, column):
        sales_records_tab = SalesRecordsSummarySheet._title
        return f"SUMIF('{sales_records_tab}'!E:E,{kpwi_cell},'{sales_records_tab}'!{column}:{column})"


class SalesRecordsSummaryWriter:

    def __init__(self, income_registry: IpBoxIncomeRegistry, projects: list[Project], year: int, path):
        self._income_registry = income_registry
        self._projects = projects
        self._year = year
        self._path = path

    def __enter__(self):
        self._workbook = Workbook()
        self._projects_sheet = ProjectsSummarySheet(self._workbook.active, self._projects)
        self._sales_records_sheet = SalesRecordsSummarySheet(self._workbook.create_sheet(), self._income_registry)
        self._results_sheet = ResultsSheet(self._workbook.create_sheet(), self._projects)
        return self

    def write_classified_sales_record(self, row: ClassifiedKpirRow):
        self._sales_records_sheet.add(row)
        self._results_sheet.add(row)

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._results_sheet.flush()
        self._workbook.save(filename=self._path)


@click.option('--year', '-y', 'years',
              help="Year in format YYYY",
              type=int,
              required=True,
              multiple=True,
              prompt="Provide year in format YYYY", )
@click.option('--timesheet', '-t', 'timesheet_dir',
              help='Timesheet reports path',
              required=True,
              default='reports/timesheet',
              type=click.Path(exists=True, file_okay=False, dir_okay=True),
              prompt="Timesheet reports directory")
@click.option('--kpir', '-k', 'kpir_dir',
              help='KPiR directory path',
              required=True,
              default='reports/record',
              type=click.Path(exists=True, file_okay=False, dir_okay=True),
              prompt="KPiR reports directory")
@click.option('--output', '-o',
              help='Record output path',
              required=True,
              default='reports/record',
              type=click.Path(exists=False, file_okay=False, dir_okay=True),
              prompt="Output directory")
@click.command()
@pass_ip_box
def summary(ip_box: IpBoxContext, years, timesheet_dir, kpir_dir, output):
    os.makedirs(output, exist_ok=True)
    income_registry = IpBoxIncomeRegistry(ip_box)
    for year in years:
        projects = ip_box.get_active_projects(year)
        for project in projects:
            click.echo(f'year: {year} project: {project.id}')
            with CsvReader(f'{timesheet_dir}/{year}/{project.id}.csv') as reader:
                for row in reader:
                    entry = AggregatedTimeEntry.from_row(row)
                    income_registry.add_ip_income(entry.date.year, entry.date.month, project, entry.ip_hours, entry.other_hours)
        with (KpirCsvReader(f'{kpir_dir}/{year}-classified.csv', skip_footer=False) as kpir_reader):
            classified_rows = [ClassifiedKpirRow(row) for row in kpir_reader]
        for row in classified_rows:
            if row.is_income:
                income_registry.add_income(row.date.year, row.date.month, row.income)
        with (SalesRecordsSummaryWriter(income_registry, projects, year,f'{output}/{year}.xlsx') as writer):
            for row in classified_rows:
                writer.write_classified_sales_record(row)
