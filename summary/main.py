#!/usr/bin/env python3

import argparse
import os
# from datetime import date, datetime, time
# from babel.dates import format_date, format_datetime, format_time
from calendar import monthrange
import csv
from decimal import Decimal

from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter


class DailyReport:

    def __init__(self, date):
        self._date = date
        self._rows = []
        self._ipbox_hours = 0
        self._other_hours = 0

    def add(self, row):
        self._rows.append(row)
        print(f"adding: {row}")
        hours = Decimal(row[2])
        if row[3] == 'True':
            self._ipbox_hours += hours
        else:
            self._other_hours += hours

    @property
    def ipbox(self):
        return self._ipbox_hours

    @property
    def other(self):
        return self._other_hours


class MonthlyReport:

    def __init__(self, reader, year, month):
        self._reader = reader
        self._year = year
        self._month = month
        self._read()


    def _read(self):
        self._report = dict()
        _, days_count = monthrange(self._year, self._month)
        for day in range(1, days_count + 1):
            date = f"{self._year}-{self._month:02d}-{day:02d}"
            self._report[date] = DailyReport(date)
        for row in self._reader:
            date = row[0]
            self._report[date].add(row[1:])

    def __getitem__(self, day):
        key = f"{self._year}-{self._month:02d}-{day:02d}"
        return self._report[key]


class ReportReader:

    def __init__(self, report_dir, projects=None, prefixes=None):
        self._projects_files = dict()
        self._report_dir = report_dir
        self._prefixes = prefixes or []
        for path in os.listdir(self._report_dir):
            if os.path.isfile(f"{self._report_dir}/{path}") and path[-4:] == '.csv':
                project_name = path[:-4]
                for prefix in self._prefixes:
                    if project_name.startswith(prefix):
                        project_name = project_name[len(prefix):]
                        break
                if not projects or project_name in projects:
                    self._projects_files[project_name] = path

    def __iter__(self):
        for file in self._projects_files.values():
            with open(f"{self._report_dir}/{file}", newline='') as csvfile:
                reader = csv.reader(csvfile, delimiter=',')
                next(reader) # skip header
                for row in reader:
                    yield row

    @property
    def existing_projects(self):
        return self._projects_files.keys()


class IPBoxWorkSheet:

    _summary = 'Sumy'

    _sheets = ['Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec', 'Lipiec', 'Sierpień', 'Wrzesień',
              'Październik', 'Listopad', 'Grudzień']

    _header = ['Dzień',	'KWIP',	'Inne',	'Łącznie', 'Podstawa obliczenia']

    def __init__(self, reports_dir, year, projects=None, prefixes=None):
        self._reports_dir = reports_dir
        self._year = year
        self._projects = projects
        self._prefixes = prefixes or []

    def _set_header(self, sheet, cell, value):
        sheet[cell] = value
        sheet[cell].font = Font(color='FF000000', bold=True)
        sheet[cell].fill = PatternFill("solid", fgColor=Color(indexed=22))

    def _generate_summary(self, sheet):
        self._set_header(sheet, 'B2', 'Pracownik')
        self._set_header(sheet, 'B3', 'Identyfikator KPWI')
        self._set_header(sheet, 'F2', 'Godzin pracy')
        self._set_header(sheet, 'G2', 'W tym KPWI')
        self._set_header(sheet, 'H2', 'Procent')
        for letter in ('B', 'C'):
            sheet.column_dimensions[letter].width = 30
        for letter in ('E', 'F', 'G', 'H'):
            sheet.column_dimensions[letter].width = 15
        for idx, month in enumerate(self._sheets):
            row = idx + 3
            sheet[f'E{row}'] = month
            sheet[f'F{row}'] = f"='{month}'!H2"
            sheet[f'G{row}'] = f"='{month}'!H3"
            sheet[f'H{row}'] = f"=G{row}/F{row}"
            sheet[f'H{row}'].number_format = '0.00%'
        sheet['E15'] = 'SUMA'
        sheet['F15'] = '=SUM(F3:F14)'
        sheet['G15'] = '=SUM(G3:G14)'
        sheet['H15'] = '=G15/F15'
        sheet[f'H15'].number_format = '0.00%'

    def _generate_month(self, sheet, month):
        reader = ReportReader(f"{self._reports_dir}/{self._year}-{month:02d}",
                              projects=self._projects, prefixes=self._prefixes)
        report = MonthlyReport(reader, self._year, month)
        for idx, header in enumerate(self._header):
            cell = f"{get_column_letter(idx + 1)}1"
            self._set_header(sheet, cell, header)
        sheet.column_dimensions[get_column_letter(len(self._header))].width = 20
        sheet.column_dimensions['G'].width = 15
        _, days_count = monthrange(self._year, month)
        for day in range(1, days_count + 1):
            row = day + 1
            sheet[f'A{row}'] = day
            sheet[f'B{row}'] = report[day].ipbox
            sheet[f'C{row}'] = report[day].other
            sheet[f'D{row}'] = f'=B{row}+C{row}'
        sheet['G2'] = 'Suma w miesiącu'
        sheet['H2'] = f"=SUM(D2:D{days_count + 1})"
        sheet['G3'] = 'w tym KPWI'
        sheet['H3'] = f"=SUM(B2:B{days_count + 1})"
        sheet['G4'] = 'procentowo KPWI'
        sheet['H4'] = "=H3/H2"
        sheet['H4'].number_format = '0.00%'

    def generate(self, report_name='karta-pracy.xlsx'):
        os.makedirs(f'{self._reports_dir}/{self._year}', exist_ok=True)
        wb = Workbook()
        summary = wb.active
        for month in range(1, len(self._sheets) + 1):
            sheet = wb.create_sheet(f"{month:02d}")
            sheet.title = self._sheets[month - 1]
            self._generate_month(sheet, month)
        summary.title = self._summary
        self._generate_summary(summary)
        wb.save(filename=f'{self._reports_dir}/{self._year}/{report_name}')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--year', help='YYYY format', required=True)
    parser.add_argument('--path', help='Reports path', required=True)
    parser.add_argument('--projects', help='IP Box compatible projects name', required=False, nargs='+')
    parser.add_argument('--prefixes', help='Reports files prefixes', required=False, nargs='*',
                        default=['Harvest ', 'Upwork '])
    args = parser.parse_args()

    os.makedirs(f'{args.path}/{args.year}', exist_ok=True)

    IPBoxWorkSheet(args.path, int(args.year), args.projects, args.prefixes).generate()
